import pandas as pd
from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook
from io import BytesIO

def clear_format(data: pd.DataFrame) -> pd.DataFrame:
   data.dropna(how="all", inplace=True)
   data.dropna(axis=1, how="all", inplace=True)
   data.reset_index(drop=True, inplace=True)
   return data

def parse_excel_file(file: UploadFile) -> dict:
   if not file.filename.endswith((".xls", ".xlsx")):
      raise HTTPException(status_code=400, detail="Invalid file format. Only .xls and .xlsx are allowed.")
   
   try:
      contents = file.file.read()
      excel_data = pd.ExcelFile(contents)
      data_frames = {}
      
      for sheet_name in excel_data.sheet_names:
         df = excel_data.parse(sheet_name)
         df = clear_format(df)
         data_frames[sheet_name] = df
      return data_frames
   
   except Exception as e:
      try:
         file.file.seek(0) 
         workbook = load_workbook(filename=BytesIO(file.file.read()))
         data_frames = {}
         
         for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            data = []

            for row in sheet.iter_rows(values_only=True):
               data.append(row)

            if not data or all(cell is None for cell in data[0]):
               continue  
            
            df = pd.DataFrame(data)

            if not df.empty:
               df.columns = df.iloc[0]
               df = df[1:]

            df = clear_format(df)
            data_frames[sheet_name] = df
         return data_frames
      
      except Exception as format_error:
         raise HTTPException(
            status_code=500,
            detail=f"Error reading Excel file: {str(format_error)}"
         )
